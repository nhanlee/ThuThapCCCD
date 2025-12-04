import os
import json
import base64
import logging
import io
from datetime import datetime
from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import pymysql
import pymysql.cursors
import traceback

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-in-production'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Database configuration
DB_HOST = "yamanote.proxy.rlwy.net"
DB_PORT = 22131
DB_USER = "root"
DB_PASS = "wIGaLEezXhTLlSShztFWktORKCeSaEGO"
DB_NAME = "railway"

MYSQL_CONFIG = {
    'host': DB_HOST,
    'port': DB_PORT,
    'user': DB_USER,
    'password': DB_PASS,
    'database': DB_NAME,
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor,
    'connect_timeout': 10
}

# Khởi tạo database
def init_db():
    connection = None
    try:
        connection = pymysql.connect(**MYSQL_CONFIG)
        logger.info(f"✅ Kết nối database thành công! Host: {DB_HOST}, Database: {DB_NAME}")
        
        with connection.cursor() as cursor:
            # Tạo bảng users
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS `users` (
                    `id` INT AUTO_INCREMENT PRIMARY KEY,
                    `username` VARCHAR(100) UNIQUE NOT NULL,
                    `fullname` VARCHAR(255),
                    `role` VARCHAR(50) DEFAULT 'user',
                    `created_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            
            # Tạo bảng id_records (đổi từ cccd_records)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS `id_records` (
                    `id` INT AUTO_INCREMENT PRIMARY KEY,
                    `cccd_moi` VARCHAR(50) UNIQUE NOT NULL,
                    `cmnd_cu` VARCHAR(50),
                    `name` VARCHAR(255),
                    `dob` DATE,
                    `gender` VARCHAR(20),
                    `address` TEXT,
                    `issue_date` DATE,
                    `phone` VARCHAR(20),
                    `user` VARCHAR(100),
                    `front_image` LONGTEXT,
                    `back_image` LONGTEXT,
                    `signature_image` LONGTEXT,
                    `created_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            
            # Kiểm tra và thêm user mẫu
            cursor.execute("SELECT COUNT(*) as count FROM users")
            result = cursor.fetchone()
            if result and result['count'] == 0:
                cursor.execute("""
                    INSERT INTO users (username, fullname, role) VALUES 
                    ('admin', 'Quản trị viên', 'admin')
                """)
                
        connection.commit()
        logger.info("✅ Database initialized successfully")
        
        # Đếm số lượng records
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) as count FROM id_records")
            result = cursor.fetchone()
            logger.info(f"✅ Số lượng records trong database: {result['count'] if result else 0}")
            
    except Exception as e:
        logger.error(f"Error initializing database: {e}")
        logger.error(traceback.format_exc())
        raise e
    finally:
        if connection:
            connection.close()

# Gọi init_db() khi khởi động
try:
    init_db()
    logger.info("✅ Database initialization completed successfully")
except Exception as e:
    logger.error(f"Could not initialize database on startup: {e}")

# Hàm kiểm tra kết nối database
def check_db_connection():
    try:
        connection = pymysql.connect(**MYSQL_CONFIG)
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            result = cursor.fetchone()
            if result and result.get('1') == 1:
                return True, "✅ Kết nối database thành công!"
        connection.close()
        return True, "✅ Kết nối database thành công!"
    except Exception as e:
        error_msg = f"Lỗi kết nối database: {str(e)}"
        logger.error(error_msg)
        return False, error_msg

# Middleware kiểm tra đăng nhập
@app.before_request
def check_login():
    # Các route không cần đăng nhập
    public_routes = ['login', 'static', 'health', 'get_user_info', 'testConnection', 'check_health', 'export_excel']
    
    if request.endpoint in public_routes:
        return
    
    if 'user_id' not in session:
        return redirect(url_for('login'))

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('index.html')
    
    data = request.get_json()
    username = data.get('username', '').strip()
    
    if not username:
        return jsonify({'success': False, 'message': 'Vui lòng nhập tên đăng nhập'})
    
    connection = None
    try:
        connection = pymysql.connect(**MYSQL_CONFIG)
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            user = cursor.fetchone()
            
            if user:
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['fullname'] = user['fullname']
                session['role'] = user['role']
                
                logger.info(f"✅ User {username} đăng nhập thành công")
                
                return jsonify({
                    'success': True,
                    'message': 'Đăng nhập thành công',
                    'user': {
                        'id': user['id'],
                        'username': user['username'],
                        'fullname': user['fullname'],
                        'role': user['role']
                    }
                })
            else:
                return jsonify({'success': False, 'message': 'Tên đăng nhập không tồn tại'})
                
    except Exception as e:
        logger.error(f"Lỗi đăng nhập: {str(e)}")
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'})
    finally:
        if connection:
            connection.close()

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/get_user_info')
def get_user_info():
    if 'user_id' not in session:
        return jsonify({'authenticated': False})
    
    return jsonify({
        'authenticated': True,
        'user': {
            'id': session.get('user_id'),
            'username': session.get('username'),
            'fullname': session.get('fullname'),
            'role': session.get('role')
        }
    })

@app.route('/saveCCCD', methods=['POST'])
def save_cccd():
    try:
        data = request.get_json()
        
        logger.info(f"📥 Nhận request saveCCCD từ user: {session.get('username')}")
        
        if not data or 'cccd_moi' not in data:
            logger.error("Thiếu số CCCD trong request")
            return jsonify({
                'success': False,
                'error': 'missing_cccd',
                'message': 'Thiếu số CCCD'
            }), 400

        cccd_number = str(data['cccd_moi']).strip()
        
        if not cccd_number:
            logger.error("Số CCCD rỗng")
            return jsonify({
                'success': False,
                'error': 'empty_cccd',
                'message': 'Số CCCD không được để trống'
            }), 400
        
        logger.info(f"🔍 Bắt đầu xử lý CCCD: {cccd_number}")
        
        # Kiểm tra trùng CCCD
        is_duplicate = check_duplicate_cccd(cccd_number)
        if is_duplicate:
            logger.warning(f"⚠️ CCCD {cccd_number} đã tồn tại")
            return jsonify({
                'success': False,
                'error': 'duplicate',
                'message': f'Số CCCD {cccd_number} đã tồn tại trong hệ thống!',
                'duplicateCCCD': cccd_number
            }), 400
        
        logger.info("✅ CCCD chưa tồn tại, tiếp tục xử lý...")

        # Lấy ảnh mặt trước, mặt sau và chữ ký
        front_base64 = data.get('front')
        back_base64 = data.get('back')
        signature_base64 = data.get('signature')
        
        if not front_base64 or not back_base64 or not signature_base64:
            logger.error("Thiếu ảnh mặt trước, mặt sau hoặc chữ ký")
            return jsonify({
                'success': False,
                'error': 'missing_images',
                'message': 'Vui lòng chụp đầy đủ ảnh mặt trước, mặt sau và chữ ký CCCD.'
            }), 400
        
        logger.info("🖼️ Bắt đầu xử lý ảnh...")

        # Lưu vào database
        connection = None
        try:
            connection = pymysql.connect(**MYSQL_CONFIG)
            
            with connection.cursor() as cursor:
                sql = '''
                    INSERT INTO id_records 
                    (cccd_moi, cmnd_cu, name, dob, gender, address, issue_date, phone, user, front_image, back_image, signature_image)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                '''
                
                # Chuyển đổi ngày tháng
                dob = None
                if data.get('dob'):
                    try:
                        dob = datetime.strptime(data['dob'], '%d/%m/%Y').date()
                    except:
                        pass
                
                issue_date = None
                if data.get('issue_date'):
                    try:
                        issue_date = datetime.strptime(data['issue_date'], '%d/%m/%Y').date()
                    except:
                        pass
                
                cursor.execute(sql, (
                    cccd_number,
                    data.get('cmnd_cu', ''),
                    data.get('name', ''),
                    dob,
                    data.get('gender', ''),
                    data.get('address', ''),
                    issue_date,
                    data.get('phone', ''),
                    session.get('username'),
                    front_base64,
                    back_base64,
                    signature_base64
                ))
            connection.commit()
            logger.info(f"✅ Đã lưu CCCD {cccd_number} vào database")
        except Exception as db_error:
            logger.error(f"Lỗi khi lưu vào database: {db_error}")
            raise
        finally:
            if connection:
                connection.close()

        logger.info(f"🎉 Hoàn tất xử lý CCCD {cccd_number}")
        
        return jsonify({
            'success': True,
            'message': 'Lưu thành công!',
            'cccd': cccd_number
        })

    except pymysql.err.IntegrityError as e:
        if 'Duplicate entry' in str(e):
            logger.warning(f"⚠️ CCCD {cccd_number} đã tồn tại (lỗi integrity)")
            return jsonify({
                'success': False,
                'error': 'duplicate',
                'message': f'Số CCCD {cccd_number} đã tồn tại trong hệ thống!',
                'duplicateCCCD': cccd_number
            }), 400
        else:
            logger.error(f"Database integrity error: {str(e)}")
            return jsonify({
                'success': False,
                'error': 'database_error',
                'message': f'Lỗi cơ sở dữ liệu: {str(e)}'
            }), 500
    except pymysql.err.OperationalError as e:
        logger.error(f"Database connection error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'connection_error',
            'message': 'Không thể kết nối đến cơ sở dữ liệu. Vui lòng thử lại sau.'
        }), 500
    except Exception as e:
        logger.error(f"Lỗi server khi saveCCCD: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'server_error',
            'message': f'Lỗi server: {str(e)}'
        }), 500

@app.route('/checkDuplicate', methods=['POST'])
def check_duplicate():
    try:
        data = request.get_json()
        cccd_number = str(data.get('cccd', '')).strip()
        
        logger.info(f"🔍 Kiểm tra trùng CCCD: {cccd_number}")
        
        is_duplicate = check_duplicate_cccd(cccd_number)
        
        return jsonify({'duplicate': is_duplicate})
    
    except Exception as e:
        logger.error(f"Lỗi kiểm tra trùng: {str(e)}")
        return jsonify({'duplicate': False})

@app.route('/get_records', methods=['GET'])
def get_records():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        search = request.args.get('search', '')
        offset = (page - 1) * limit
        
        connection = None
        try:
            connection = pymysql.connect(**MYSQL_CONFIG)
            
            with connection.cursor() as cursor:
                # Đếm tổng số bản ghi
                count_sql = "SELECT COUNT(*) as total FROM id_records"
                count_params = []
                
                if search:
                    count_sql += " WHERE cccd_moi LIKE %s OR name LIKE %s OR phone LIKE %s"
                    search_term = f"%{search}%"
                    count_params = [search_term, search_term, search_term]
                
                cursor.execute(count_sql, count_params)
                count_result = cursor.fetchone()
                total = count_result['total'] if count_result else 0
                
                # Lấy dữ liệu
                sql = """
                    SELECT id, cccd_moi, cmnd_cu, name, dob, gender, address, issue_date, phone, user, created_at 
                    FROM id_records 
                """
                params = []
                
                if search:
                    sql += " WHERE cccd_moi LIKE %s OR name LIKE %s OR phone LIKE %s"
                    search_term = f"%{search}%"
                    params = [search_term, search_term, search_term]
                
                sql += " ORDER BY created_at DESC LIMIT %s OFFSET %s"
                params.extend([limit, offset])
                
                cursor.execute(sql, params)
                records = cursor.fetchall()
                
                # Format ngày tháng
                for record in records:
                    if record['dob']:
                        record['dob'] = record['dob'].strftime('%d/%m/%Y')
                    if record['issue_date']:
                        record['issue_date'] = record['issue_date'].strftime('%d/%m/%Y')
                    if record['created_at']:
                        record['created_at'] = record['created_at'].strftime('%d/%m/%Y %H:%M')
                
                return jsonify({
                    'success': True,
                    'records': records,
                    'total': total,
                    'page': page,
                    'total_pages': (total + limit - 1) // limit
                })
                
        except Exception as e:
            logger.error(f"Lỗi khi lấy records: {str(e)}")
            return jsonify({'success': False, 'message': str(e)})
        finally:
            if connection:
                connection.close()
                
    except Exception as e:
        logger.error(f"Lỗi get_records: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_record_detail/<int:record_id>', methods=['GET'])
def get_record_detail(record_id):
    try:
        connection = None
        try:
            connection = pymysql.connect(**MYSQL_CONFIG)
            
            with connection.cursor() as cursor:
                sql = """
                    SELECT * FROM id_records WHERE id = %s
                """
                cursor.execute(sql, (record_id,))
                record = cursor.fetchone()
                
                if not record:
                    return jsonify({'success': False, 'message': 'Không tìm thấy bản ghi'})
                
                # Format ngày tháng
                if record['dob']:
                    record['dob'] = record['dob'].strftime('%d/%m/%Y')
                if record['issue_date']:
                    record['issue_date'] = record['issue_date'].strftime('%d/%m/%Y')
                if record['created_at']:
                    record['created_at'] = record['created_at'].strftime('%d/%m/%Y %H:%M')
                
                return jsonify({
                    'success': True,
                    'record': record
                })
                
        except Exception as e:
            logger.error(f"Lỗi khi lấy chi tiết record: {str(e)}")
            return jsonify({'success': False, 'message': str(e)})
        finally:
            if connection:
                connection.close()
                
    except Exception as e:
        logger.error(f"Lỗi get_record_detail: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_all_images/<int:record_id>', methods=['GET'])
def get_all_images(record_id):
    try:
        connection = None
        try:
            connection = pymysql.connect(**MYSQL_CONFIG)
            
            with connection.cursor() as cursor:
                sql = """
                    SELECT front_image, back_image, signature_image 
                    FROM id_records WHERE id = %s
                """
                cursor.execute(sql, (record_id,))
                record = cursor.fetchone()
                
                if not record:
                    return jsonify({'success': False, 'message': 'Không tìm thấy bản ghi'})
                
                return jsonify({
                    'success': True,
                    'images': {
                        'front': record['front_image'],
                        'back': record['back_image'],
                        'signature': record['signature_image']
                    }
                })
                
        except Exception as e:
            logger.error(f"Lỗi khi lấy ảnh: {str(e)}")
            return jsonify({'success': False, 'message': str(e)})
        finally:
            if connection:
                connection.close()
                
    except Exception as e:
        logger.error(f"Lỗi get_all_images: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

def check_duplicate_cccd(cccd_number):
    """Kiểm tra CCCD có trùng không"""
    connection = None
    try:
        connection = pymysql.connect(**MYSQL_CONFIG)
        with connection.cursor() as cursor:
            cursor.execute('SELECT COUNT(*) as count FROM id_records WHERE cccd_moi = %s', (cccd_number,))
            result = cursor.fetchone()
            count = result['count'] if result else 0
        return count > 0
    
    except Exception as e:
        logger.error(f"Lỗi kiểm tra trùng CCCD: {str(e)}")
        return False
    finally:
        if connection:
            connection.close()

@app.route('/export_excel', methods=['GET'])
def export_excel():
    try:
        # Tạo workbook mới
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách CCCD"
        
        # Tạo header
        headers = ['STT', 'Số CCCD', 'Số CMND cũ', 'Họ và tên', 'Ngày sinh', 
                   'Giới tính', 'Địa chỉ', 'Ngày cấp', 'Số điện thoại', 
                   'Người nhập', 'Ngày tạo']
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = Font(bold=True)
            ws[f'{col_letter}1'].alignment = Alignment(horizontal='center')
        
        # Lấy dữ liệu từ database
        connection = None
        try:
            connection = pymysql.connect(**MYSQL_CONFIG)
            with connection.cursor() as cursor:
                sql = """
                    SELECT cccd_moi, cmnd_cu, name, dob, gender, address, issue_date, phone, user, created_at 
                    FROM id_records 
                    ORDER BY created_at DESC
                """
                cursor.execute(sql)
                records = cursor.fetchall()
                
                # Thêm dữ liệu vào excel
                for row_num, record in enumerate(records, 2):
                    ws[f'A{row_num}'] = row_num - 1
                    ws[f'B{row_num}'] = record['cccd_moi']
                    ws[f'C{row_num}'] = record['cmnd_cu'] or ''
                    ws[f'D{row_num}'] = record['name']
                    ws[f'E{row_num}'] = record['dob'].strftime('%d/%m/%Y') if record['dob'] else ''
                    ws[f'F{row_num}'] = record['gender']
                    ws[f'G{row_num}'] = record['address']
                    ws[f'H{row_num}'] = record['issue_date'].strftime('%d/%m/%Y') if record['issue_date'] else ''
                    ws[f'I{row_num}'] = record['phone'] or ''
                    ws[f'J{row_num}'] = record['user']
                    ws[f'K{row_num}'] = record['created_at'].strftime('%d/%m/%Y %H:%M')
                
                # Điều chỉnh độ rộng cột
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Lưu workbook vào buffer
                from io import BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                # Trả về file excel
                return send_file(
                    buffer,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'danh_sach_cccd_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                )
                
        except Exception as e:
            logger.error(f"Lỗi khi export excel: {str(e)}")
            return jsonify({'success': False, 'message': str(e)})
        finally:
            if connection:
                connection.close()
                
    except Exception as e:
        logger.error(f"Lỗi export_excel: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/testConnection', methods=['GET'])
def test_connection():
    try:
        # Kiểm tra kết nối database
        db_success, db_message = check_db_connection()
        
        # Trạng thái tổng thể
        if db_success:
            status = 'ok'
            message = 'Tất cả hệ thống hoạt động bình thường'
        else:
            status = 'warning'
            message = 'Một số hệ thống có vấn đề'
        
        return jsonify({
            'status': status,
            'message': message,
            'database': {
                'success': db_success,
                'message': db_message,
                'host': DB_HOST,
                'port': DB_PORT,
                'database': DB_NAME
            },
            'scanner': {
                'success': True,
                'message': '✅ QR Scanner sẵn sàng',
                'technology': 'jsQR'
            }
        })
    
    except Exception as e:
        logger.error(f"Lỗi testConnection: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/health')
def health_check():
    try:
        db_success, db_message = check_db_connection()
        
        return jsonify({
            'status': 'healthy' if db_success else 'unhealthy', 
            'database': db_message,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/check_health')
def check_health():
    """Endpoint đơn giản để kiểm tra sức khỏe hệ thống"""
    try:
        connection = pymysql.connect(**MYSQL_CONFIG)
        connection.close()
        return jsonify({'status': 'healthy', 'message': 'Hệ thống hoạt động bình thường'})
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'message': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    logger.info(f"🚀 Ứng dụng khởi động trên cổng {port}")
    logger.info(f"📊 Database: {DB_HOST}:{DB_PORT}/{DB_NAME}")
    app.run(host='0.0.0.0', port=port, debug=False)